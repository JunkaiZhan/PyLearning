#!/usr/bin/python3
# -*- coding: UTF-8 -*-
# Author: zhanjunkai

import argparse
import os
import sys
import re
import logging
import openpyxl
import subprocess

# =============================================
# Function    : sh(cmd)
# Description : do cmd with bash shell
# ============================================= 
def sh(cmd):
    subprocess.call(cmd, shell=True)

# ====================================================================
# Function    : excel_letter2int(letter)
# Description : transform the letter to int number for column of excel
# ====================================================================
def excel_letter2int(letter):
    return openpyxl.utils.column_index_from_string(letter)


# =====================================================
# Function    : get_cell_from_sheet(sheet, row, column)
# Description : get the cell value from the work sheet
# ===================================================== 
def get_cell_from_sheet(sheet, row, column):
    cell_value = sheet.cell(row=row, column=column).value
    return cell_value


# ===========================================================================
# Function    : get_row_from_sheet(sheet, row_index, column_start, column_end)
# Description : get the list of values in row_index from column_start
#               to column_end
# ===========================================================================
def get_row_from_sheet(sheet, row_index, column_start, column_end):
    row_list = []
    for _ in range(column_start, column_end+1):
        cell_value = get_cell_from_sheet(sheet, row_index, _)
        row_list.append(cell_value)
    # check_empty(row_list)
    return row_list

# ====================================================================
# Function    : get_index_in_row_header(sheet, item_name, header_pos = 1)
# Description : get the index of the item in the header
# ====================================================================
def get_index_in_row_header (sheet, item_name, header_pos = 1):
    index    = -1
    max_col  = sheet.max_column
    row_list = get_row_from_sheet(sheet, header_pos, 1, max_col)
    for _ in range(0, len(row_list)):
        if (item_name == row_list[_]):
            index = _ + 1
            return index
    print("The item: " + item_name + " is not found in the header, please check the item name or header_pos")


# ====================================================================
# Function    : get_index_in_col_header(sheet, item_name, header_pos = 1)
# Description : get the index of the item in the header
# ====================================================================
def get_index_in_col_header (sheet, item_name, header_pos = 1):
    index    = -1
    max_row  = sheet.max_row
    col_list = get_row_from_sheet(sheet, header_pos, max_row, 1)
    for _ in range(0, len(col_list)):
        if (item_name == col_list[_]):
            index = _ + 1
            return index
    print("The item: " + item_name + " is not found in the header, please check the item name or header_pos")


# ====================================================================
# Function   : create_hier_dict(hier_file_path)
# Description: Create the hierachy dictionary
# ====================================================================
def create_hier_dict(hier_file_path):

    hier_dict = {}

    # Get hierarchy macro file handle and read lines
    with open(hier_file_path, 'r') as hier_handle:
        hier_lines  = hier_handle.readlines()

    # Scan the hierarchy hier_lines and create dict item
    for line in hier_lines:
        if "`define" in line and not "//" in line:
            temp_list = line.split()
            hier_dict[temp_list[1]] = temp_list[2]

    # Return
    return hier_dict


# ====================================================================
# Function   : create_vip_dict(vip_path_file, hier_dict)
# Description: Create the VIP path list
# ====================================================================
def create_vip_list(vip_path_file, hier_dict):

    mst_vip_list = []
    slv_vip_list = []
    ext_vip_list = []

    # Get hierarchy macro file handle
    wb = openpyxl.load_workbook(vip_path_file)
    master_sheet = wb["master_port_config"]
    slave_sheet  = wb["slave_port_config"]
    extra_sheet  = wb["extra_vip_config"]

    print("Scan the vip_lines for master sheet -----------------------------------")
    cate_index = get_index_in_row_header(master_sheet, "Config Category")
    inst_index = get_index_in_row_header(master_sheet, "Port Instance Name")
    mode_index = get_index_in_row_header(master_sheet, "Work Mode")
    cate       = cate_index - 1
    inst       = inst_index - 1
    mode       = mode_index - 1
    start_row  = 3
    for r in range(start_row, master_sheet.max_row + 1):
        row = get_row_from_sheet(master_sheet, r, 1, max(cate_index, inst_index, mode_index))
        logging.debug(str(row))
        if row[cate] is None or row[inst] is None:
            print("\033[1;35mWARNING: The row", r, "in master sheet is invalid.\033[0m")
            logging.debug("\033[1;35mBecause the Config Category is invalid or Port Instance Name is empty.\033[0m")
        elif ("AXI" in row[cate] or "AHB" in row[cate]) and not row[inst] is None: # if the row is valid
            path = re.search(r'[\w\+]+', row[inst], re.I|re.M).group()
            print("\033[1;34mINFO: The origin path in row", r, "in master sheet is >>>", path, "\033[0m")
            path_list = path.split("+")
            if len(path_list) > 1:
                if path_list[0][2:] in hier_dict.keys(): # Use [2:] to remove the U_ prefix
                    path = hier_dict[path_list[0][2:]] + "." + ".".join(i for i in path_list[1:])
                else:
                    path = ".".join(i for i in path_list)
            else:
                if path_list[0][2:] in hier_dict.keys():
                    path = hier_dict[path_list[0][2:]]
            print("\033[1;32mINFO: The new path in row", r, "in master sheet is >>>", path, "\033[0m")
            mst_vip_list.append(path)
        else:
            print("\033[1;35mWARNING: The row", r, "in master sheet is invalid.\033[0m")
            logging.debug("\033[1;35mPlease check the excel file and try again.\033[0m")
        
    print("Scan the vip_lines for slave sheet -----------------------------------")
    cate_index = get_index_in_row_header(slave_sheet, "Config Category")
    inst_index = get_index_in_row_header(slave_sheet, "Port Instance Name")
    mode_index = get_index_in_row_header(slave_sheet, "Work Mode")
    cate       = cate_index - 1
    inst       = inst_index - 1
    mode       = mode_index - 1
    start_row  = 3
    for r in range(start_row, slave_sheet.max_row + 1):
        row = get_row_from_sheet(slave_sheet, r, 1, max(cate_index, inst_index, mode_index))
        logging.debug(str(row))
        if row[cate] is None or row[inst] is None:
            print("\033[1;35mWARNING: The row", r, "in slave sheet is invalid.\033[0m")
            logging.debug("\033[1;35mBecause the Config Category is invalid or Port Instance Name is empty.\033[0m")
        elif ("AXI" in row[cate] or "AHB" in row[cate] or "APB" in row[cate]) and not row[inst] is None: # if the row is valid
            path = re.search(r'[\w\+]+', row[inst], re.I|re.M).group()
            print("\033[1;34mINFO: The origin path in row", r, "in slave sheet is >>>", path, "\033[0m")
            path_list = path.split("+")
            if len(path_list) > 1:
                if path_list[0][2:] in hier_dict.keys(): # Use [2:] to remove the U_ prefix
                    path = hier_dict[path_list[0][2:]] + "." + ".".join(i for i in path_list[1:])
                else:
                    path = ".".join(i for i in path_list)
            else:
                if path_list[0][2:] in hier_dict.keys():
                    path = hier_dict[path_list[0][2:]]
            print("\033[1;32mINFO: The new path in row", r, "in slave sheet is >>>", path, "\033[0m")
            slv_vip_list.append(path)
        else:
            print("\033[1;35mWARNING: The row", r, "in slave sheet is invalid.\033[0m")
            logging.debug("\033[1;35mPlease check the excel file and try again.\033[0m")

    print("Scan the vip_lines for extra sheet -----------------------------------")
    cate_index = get_index_in_row_header(extra_sheet, "VIP Name")
    inst_index = get_index_in_row_header(extra_sheet, "Port Instance Name")
    mode_index = get_index_in_row_header(extra_sheet, "Work Mode")
    cate       = cate_index - 1
    inst       = inst_index - 1
    mode       = mode_index - 1
    start_row  = 3
    for r in range(start_row, extra_sheet.max_row + 1):
        row = get_row_from_sheet(extra_sheet, r, 1, max(cate_index, inst_index, mode_index))
        logging.debug(str(row))
        if row[cate] is None or row[inst] is None:
            print("\033[1;35mWARNING: The row", r, "in extra sheet is invalid.\033[0m")
            logging.debug("\033[1;35mBecause the Config Category is invalid or Port Instance Name is empty.\033[0m")
        elif ("axi" in row[cate] or "ahb" in row[cate] or "apb" in row[cate]) and not row[inst] is None: # if the row is valid
            path = re.search(r'[\w\+]+', row[inst], re.I|re.M).group()
            print("\033[1;34mINFO: The origin path in row", r, "in extra sheet is >>>", path, "\033[0m")
            path_list = path.split("+")
            if len(path_list) > 1:
                if path_list[0][2:] in hier_dict.keys(): # Use [2:] to remove the U_ prefix
                    path = hier_dict[path_list[0][2:]] + "." + ".".join(i for i in path_list[1:])
                else:
                    path = ".".join(i for i in path_list)
            else:
                if path_list[0][2:] in hier_dict.keys():
                    path = hier_dict[path_list[0][2:]]
            print("\033[1;32mINFO: The new path in row", r, "in extra sheet is >>>", path, "\033[0m")
            ext_vip_list.append(path)
        else:
            print("\033[1;35mWARNING: The row", r, "in extra sheet is invalid.\033[0m")
            logging.debug("\033[1;35mPlease check the excel file and try again.\033[0m")

    return mst_vip_list, slv_vip_list, ext_vip_list

# ====================================================================
# Function   : vip_path_check(kwargs)
# Description: 
# ====================================================================
def vip_path_check(kwargs):

    # Get dict: macro -> full_path
    hier_dict = create_hier_dict(kwargs["macro"])
    
    # Get vip path list from excel
    mst_vip_list, slv_vip_list, ext_vip_list = create_vip_list(kwargs["vip"], hier_dict)
    print(mst_vip_list)
    print(slv_vip_list)
    print(ext_vip_list)


# Main -----------------------------------------------
if __name__ == "__main__":

    logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
    logging.disable(logging.DEBUG)
    
    options = argparse.ArgumentParser()

    # get arguments from command
    options.add_argument("-m", "--macro", type = str, required = True)
    options.add_argument("-v", "--vip", type = str, required = True)

    # transform the args from namespace object to dict
    args = options.parse_args()
    opts_dict = vars(args)

    # print the arguments we get 
    # for key, value in opts_dict.items():
    #     logging.debug("argument: {0} 's value is {1}".format(key, value))    

    vip_path_check(opts_dict)